"""
IP India E-Register Scraper — Application Number Search
========================================================
Scrapes trademark data from https://tmrsearch.ipindia.gov.in/estatus/
Uses application numbers from Excel input to search the E-Register portal.

Flow:
  1. Opens Chrome — user manually logs in with OTP (one-time)
  2. Navigates to Trade Mark Application/Registered Mark → National/IRDI Number
  3. For each application number: fills form, auto-solves CAPTCHA, extracts data
  4. Saves results to CSV with checkpoint support

Usage:
    python eregister_scraper.py

Input:  PCPB_Input.xlsx  (columns: "Country", "Application Number")
Output: eregister_output.csv
Resume: eregister_checkpoint.txt
"""

import csv
import json
import os
import random
import re
import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass
import time
import traceback
import urllib.request

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, StaleElementReferenceException
)

# ── Config ───────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH     = os.path.join(BASE_DIR, "PCPB_Input.xlsx")
OUTPUT_CSV     = os.path.join(BASE_DIR, "eregister_output.csv")
CHECKPOINT     = os.path.join(BASE_DIR, "eregister_checkpoint.txt")
IMAGE_DIR      = os.path.join(BASE_DIR, "tm_images")

BASE_URL       = "https://tmrsearch.ipindia.gov.in/estatus"
LOGIN_URL      = f"{BASE_URL}/OTP/index"
SELECT_URL     = f"{BASE_URL}/TradeMarkApplication/eregister"
SEARCH_URL     = f"{BASE_URL}/TradeMarkApplication/ViewRegistered"
RESULTS_URL    = f"{BASE_URL}/RegisteredTM"

WAIT_TIMEOUT       = 20
BETWEEN_SEARCH_MIN = 3    # Random delay between searches (seconds)
BETWEEN_SEARCH_MAX = 5
MAX_CAPTCHA_RETRIES = 5
BATCH_SIZE         = 50   # Take a break every N searches
BATCH_BREAK_MIN    = 30   # Coffee break duration (seconds)
BATCH_BREAK_MAX    = 60
# ─────────────────────────────────────────────────────────────────────────────

CSV_HEADERS = [
    "Search App Number", "Country",
    "Status", "Sub Status",
    "Trade Mark No.", "Date of Application", "Class", "Filing Mode",
    "Trade Mark", "TM Type", "User Detail",
    "Publication Details", "Valid Upto/ Renewed Upto",
    "Proprietor Name", "Image URL",
]

# Word-to-number map for CAPTCHA positional instructions
WORD_TO_POS = {
    "first": 1, "second": 2, "third": 3, "fourth": 4,
    "fifth": 5, "sixth": 6, "seventh": 7, "eighth": 8,
    "ninth": 9, "tenth": 10, "last": -1,
    "1st": 1, "2nd": 2, "3rd": 3, "4th": 4,
    "5th": 5, "6th": 6, "7th": 7, "8th": 8,
}


# ─── CAPTCHA Solver (Universal) ─────────────────────────────────────────────

def solve_captcha(driver):
    """
    Solve the E-Register CAPTCHA programmatically.

    Handles multiple CAPTCHA types:
      Type 1 — Positional: "Enter the third number in" → "9 5 3 3 = ?" → answer: 3
      Type 2 — Addition:   "7 + 3 = ?" → answer: 10
      Type 3 — Subtraction: "9 - 4 = ?" → answer: 5
      Type 4 — Multiplication: "6 * 3 = ?" or "6 x 3 = ?" → answer: 18
      Type 5 — Mixed:  instruction says "sum", "add", "multiply", "subtract", etc.
    """
    try:
        # Get the instruction text (e.g., "Enter the third number in")
        instruction = driver.execute_script("""
            var el = document.getElementById('captchatext');
            return el ? el.textContent.trim() : '';
        """)

        # Get the numbers/expression string (e.g., "9 5 3 3" or "7 + 3")
        numbers_str = driver.execute_script("""
            var el = document.getElementById('CaptchModel_CaptchaNumbers');
            return el ? el.value.trim() : '';
        """)

        if not numbers_str:
            print(f"    [CAPTCHA] Could not read CAPTCHA data")
            return None

        # If instruction is also empty, try to read from other elements
        if not instruction:
            instruction = driver.execute_script("""
                var labels = document.querySelectorAll('label, span, div');
                for (var i = 0; i < labels.length; i++) {
                    var t = labels[i].textContent.trim().toLowerCase();
                    if (t.indexOf('enter') > -1 && (t.indexOf('number') > -1 || t.indexOf('answer') > -1 || t.indexOf('result') > -1 || t.indexOf('captcha') > -1)) {
                        return labels[i].textContent.trim();
                    }
                }
                return '';
            """) or ""

        print(f"    [CAPTCHA] Instruction: {instruction}")
        print(f"    [CAPTCHA] Expression: {numbers_str}")

        # Clean the expression — remove trailing "= ?" or "=?"
        expr_clean = re.sub(r'\s*=\s*\??\s*$', '', numbers_str).strip()

        answer = None

        # ── Strategy 1: Check if it's a math expression (contains +, -, *, x, ×) ──
        if re.search(r'[+\-*/×x]', expr_clean, re.IGNORECASE):
            answer = _solve_math_expression(expr_clean)
            if answer is not None:
                print(f"    [CAPTCHA] Math answer: {answer}")
                return str(answer)

        # ── Strategy 2: Check if instruction asks for a position ──
        instruction_lower = instruction.lower()
        position = _parse_position(instruction_lower)

        if position is not None:
            # Parse individual numbers from the expression
            numbers = re.findall(r'\d+', expr_clean)
            if numbers:
                if position == -1:  # "last"
                    answer = numbers[-1]
                else:
                    idx = position - 1
                    if 0 <= idx < len(numbers):
                        answer = numbers[idx]

                if answer is not None:
                    print(f"    [CAPTCHA] Position answer: {answer}")
                    return str(answer)

        # ── Strategy 3: Check if instruction mentions a math operation ──
        if any(w in instruction_lower for w in ['add', 'sum', 'plus', 'total']):
            numbers = [int(n) for n in re.findall(r'\d+', expr_clean)]
            if numbers:
                answer = sum(numbers)
                print(f"    [CAPTCHA] Sum answer: {answer}")
                return str(answer)

        if any(w in instruction_lower for w in ['subtract', 'minus', 'difference']):
            numbers = [int(n) for n in re.findall(r'\d+', expr_clean)]
            if len(numbers) >= 2:
                answer = numbers[0] - numbers[1]
                print(f"    [CAPTCHA] Subtract answer: {answer}")
                return str(answer)

        if any(w in instruction_lower for w in ['multiply', 'product', 'times']):
            numbers = [int(n) for n in re.findall(r'\d+', expr_clean)]
            if len(numbers) >= 2:
                result = 1
                for n in numbers:
                    result *= n
                answer = result
                print(f"    [CAPTCHA] Multiply answer: {answer}")
                return str(answer)

        # ── Strategy 4: Fallback — try eval on the cleaned expression ──
        try:
            # Replace 'x' and '×' with '*' for multiplication
            eval_expr = expr_clean.replace('×', '*').replace('x', '*').replace('X', '*')
            # Only allow digits, spaces, and math operators
            if re.match(r'^[\d\s+\-*/().]+$', eval_expr):
                answer = int(eval(eval_expr))
                print(f"    [CAPTCHA] Eval answer: {answer}")
                return str(answer)
        except Exception:
            pass

        # ── Strategy 5: Last resort — just return the last number ──
        numbers = re.findall(r'\d+', expr_clean)
        if numbers:
            answer = numbers[-1]
            print(f"    [CAPTCHA] Fallback (last number): {answer}")
            return str(answer)

        print(f"    [CAPTCHA] Could not solve: {instruction} | {numbers_str}")
        return None

    except Exception as e:
        print(f"    [CAPTCHA] Solver error: {e}")
        return None


def _solve_math_expression(expr):
    """Solve a simple math expression like '7 + 3' or '9 - 4' or '6 x 3'."""
    try:
        # Normalize operators
        expr = expr.replace('×', '*').replace('x', '*').replace('X', '*')
        # Remove any non-math characters
        expr = re.sub(r'[^\d+\-*/().\s]', '', expr).strip()
        if not expr:
            return None
        # Safety check — only allow simple math
        if re.match(r'^[\d\s+\-*/().]+$', expr):
            return int(eval(expr))
    except Exception:
        pass
    return None


def _parse_position(instruction_lower):
    """Parse a positional word from the CAPTCHA instruction. Returns 1-based index or -1 for 'last'."""
    for word, pos in WORD_TO_POS.items():
        if word in instruction_lower:
            return pos
    # Try numeric like "enter the 3 number" or "enter number 2"
    match = re.search(r'(\d+)\s*(?:st|nd|rd|th)?\s*(?:number|digit|num)', instruction_lower)
    if match:
        return int(match.group(1))
    return None


# ─── Excel Loading ────────────────────────────────────────────────────────────

def load_search_tasks():
    """Load application numbers from the input Excel file."""
    if not os.path.exists(EXCEL_PATH):
        sys.exit(f"ERROR: Excel file not found: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    try:
        app_col = headers.index("Application Number")
    except ValueError:
        sys.exit("ERROR: 'Application Number' column not found in Excel")

    country_col = None
    if "Country" in headers:
        country_col = headers.index("Country")

    tasks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        app_num = row[app_col]
        if app_num is None:
            continue
        app_str = str(app_num).strip()
        if not app_str or app_str.lower() == "none":
            continue

        country = ""
        if country_col is not None and row[country_col]:
            country = str(row[country_col]).strip()

        tasks.append((app_str, country))

    print(f"[EXCEL] Loaded {len(tasks)} application numbers from {EXCEL_PATH}")
    return tasks


# ─── Checkpoint ───────────────────────────────────────────────────────────────

def load_checkpoint():
    done = set()
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    done.add(line)
    return done

def mark_done(app_num):
    with open(CHECKPOINT, "a", encoding="utf-8") as f:
        f.write(app_num + "\n")


# ─── CSV ──────────────────────────────────────────────────────────────────────

def init_csv():
    if not os.path.exists(OUTPUT_CSV):
        with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
            csv.DictWriter(f, fieldnames=CSV_HEADERS).writeheader()
        print(f"[CSV] Created: {OUTPUT_CSV}")
    else:
        print(f"[CSV] Will append to existing: {OUTPUT_CSV}")

def append_rows(rows):
    if not rows:
        return
    with open(OUTPUT_CSV, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CSV_HEADERS, extrasaction="ignore")
        for row in rows:
            w.writerow(row)


# ─── Browser ─────────────────────────────────────────────────────────────────

def init_driver():
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-extensions")
    # Rotate User-Agent to look like a regular browser
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
    ]
    options.add_argument(f"--user-agent={random.choice(user_agents)}")
    options.page_load_strategy = 'eager'
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(options=options)
    # Remove webdriver detection flags
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': """
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
        window.chrome = {runtime: {}};
    """})
    driver.set_script_timeout(20)
    driver.implicitly_wait(2)
    return driver


def human_delay():
    """Random delay between actions to mimic human behavior."""
    delay = random.uniform(BETWEEN_SEARCH_MIN, BETWEEN_SEARCH_MAX)
    time.sleep(delay)


def wait_for_manual_login(driver):
    """
    Navigate to E-Register and wait for user to complete OTP login.
    Returns once the user is on the Home/Index page (logged in).
    """
    driver.get(LOGIN_URL)
    print("\n" + "=" * 65)
    print("  MANUAL LOGIN REQUIRED")
    print("=" * 65)
    print("  1. Enter your Email/Phone number")
    print("  2. Click 'Generate OTP'")
    print("  3. Enter the OTP received")
    print("  4. Click 'Login'")
    print("=" * 65)
    print("\n  Waiting for login... (checking every 2 seconds)")

    while True:
        try:
            current_url = driver.current_url
            # Check if we're past the login page
            if "/Home/Index" in current_url or "/Home" in current_url:
                print("  ✓ Login successful! Detected home page.")
                return True
            if "/TradeMarkApplication" in current_url:
                print("  ✓ Login successful! Already on search page.")
                return True
            if "/RegisteredTM" in current_url:
                print("  ✓ Login successful! Already on results page.")
                return True
        except Exception:
            pass
        time.sleep(2)


def is_international(country):
    """Check if this entry needs International Registration Number search."""
    if not country:
        return False
    return country.strip().lower() not in ("india", "")


def check_session_alive(driver):
    """Check if the session is still valid. Returns True if logged in, False if expired."""
    try:
        page_text = driver.execute_script("return document.body ? document.body.innerText.substring(0, 500) : '';")
        page_src = driver.execute_script("return document.body ? document.body.innerHTML.substring(0, 1000) : '';")
        current_url = driver.current_url

        # Detect session expiry signs
        if '"code":"401"' in page_text or '"message":"Unauthorized"' in page_text or "Unauthorized" in page_text:
            return False
        if any(x in page_src for x in ['"code":"401"', '"message":"Unauthorized"']):
            return False
        if "/OTP/" in current_url or "login" in current_url.lower():
            return False
        return True
    except Exception:
        return True  # Can't check, assume alive


def handle_session_expiry(driver):
    """Detect session expiry and prompt user to re-login."""
    print("\n" + "!" * 65)
    print("  SESSION EXPIRED — Please login again!")
    print("!" * 65)

    # Navigate to the login page
    driver.get(LOGIN_URL)
    time.sleep(1)

    print("  Waiting for re-login... (complete OTP login in the browser)")

    while True:
        try:
            current_url = driver.current_url
            if "/Home/Index" in current_url or "/Home" in current_url:
                print("  ✓ Re-login successful!")
                return True
            if "/TradeMarkApplication" in current_url:
                print("  ✓ Re-login successful!")
                return True
        except Exception:
            pass
        time.sleep(2)


def navigate_to_search_form(driver, international=False):
    """Navigate to the search form and select the correct radio button."""
    search_type = "International" if international else "National/IRDI"
    try:
        driver.get(SELECT_URL)
        time.sleep(0.8)

        # Check for session expiry
        if not check_session_alive(driver):
            handle_session_expiry(driver)
            # After re-login, navigate again
            driver.get(SELECT_URL)
            time.sleep(0.8)

        if international:
            # Click "International Registration Number" radio button
            radio_clicked = driver.execute_script("""
                var radios = document.querySelectorAll('input[type="radio"]');
                for (var i = 0; i < radios.length; i++) {
                    var label = radios[i].parentElement ? radios[i].parentElement.textContent : '';
                    if (label.indexOf('International') > -1) {
                        radios[i].click();
                        return true;
                    }
                }
                // Try by value
                for (var j = 0; j < radios.length; j++) {
                    if (radios[j].value === 'I' || radios[j].value === 'international') {
                        radios[j].click();
                        return true;
                    }
                }
                // Click the second radio
                if (radios.length > 1) { radios[1].click(); return true; }
                return false;
            """)
        else:
            # Click "National/IRDI Number" radio button
            radio_clicked = driver.execute_script("""
                var radios = document.querySelectorAll('input[type="radio"]');
                for (var i = 0; i < radios.length; i++) {
                    var label = radios[i].parentElement ? radios[i].parentElement.textContent : '';
                    if (label.indexOf('National') > -1 || label.indexOf('IRDI') > -1) {
                        radios[i].click();
                        return true;
                    }
                }
                // Try by value
                for (var j = 0; j < radios.length; j++) {
                    if (radios[j].value === 'N' || radios[j].value === 'national') {
                        radios[j].click();
                        return true;
                    }
                }
                // Click the first radio
                if (radios.length > 0) { radios[0].click(); return true; }
                return false;
            """)

        if not radio_clicked:
            print(f"  [WARN] Could not click {search_type} radio button")
            return False

        time.sleep(0.8)

        # Verify the search form appeared
        app_field = driver.find_elements(By.ID, "ApplicationNumber")
        if app_field:
            print(f"  [OK] {search_type} search form is ready.")
            return True
        else:
            # Try direct URL as fallback (only works for National)
            if not international:
                driver.get(SEARCH_URL)
                time.sleep(0.8)
                app_field = driver.find_elements(By.ID, "ApplicationNumber")
                if app_field:
                    print(f"  [OK] {search_type} search form is ready (via direct URL).")
                    return True
            print(f"  [WARN] {search_type} search form not found")
            return False

    except Exception as e:
        print(f"  [ERROR] Navigation failed: {e}")
        return False


# ─── Search & Extract ─────────────────────────────────────────────────────────

def go_back_to_form(driver):
    """
    Click the Back button on the results page to return to the search form.
    Returns True if form is ready, False if we need full re-navigation.
    """
    try:
        # Try clicking the Back button (it's an <input type="button"> or <a>)
        clicked = driver.execute_script("""
            // Look for Back button
            var btns = document.querySelectorAll('input[type="button"], input[type="submit"], button, a');
            for (var i = 0; i < btns.length; i++) {
                var txt = (btns[i].value || btns[i].textContent || '').trim().toLowerCase();
                if (txt === 'back') {
                    btns[i].click();
                    return true;
                }
            }
            // Try browser back
            window.history.back();
            return true;
        """)

        if clicked:
            # Wait for the form to appear
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "ApplicationNumber"))
                )
                return True
            except TimeoutException:
                return False
        return False
    except Exception:
        return False

def fill_and_submit(driver, app_number, international=False):
    """Fill the search form, solve CAPTCHA, and click View."""
    for attempt in range(MAX_CAPTCHA_RETRIES):
        try:
            if attempt > 0:
                # Check if we're still on the form page
                still_on_form = driver.find_elements(By.ID, "ApplicationNumber")
                if still_on_form:
                    # Just refresh the CAPTCHA — no need to re-navigate
                    driver.execute_script("""
                        var links = document.querySelectorAll('a');
                        for (var i = 0; i < links.length; i++) {
                            var oc = links[i].getAttribute('onclick') || '';
                            if (oc.indexOf('loadCaptcha') > -1 || oc.indexOf('Captcha') > -1) {
                                links[i].click(); break;
                            }
                        }
                    """)
                    time.sleep(0.5)
                else:
                    # Lost the form — full re-navigation needed
                    navigate_to_search_form(driver, international=international)

            # Fill application number
            app_input = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.ID, "ApplicationNumber"))
            )
            app_input.clear()
            app_input.send_keys(app_number)

            # Solve CAPTCHA
            answer = solve_captcha(driver)
            if answer is None:
                print(f"    [CAPTCHA] Could not solve, refreshing ({attempt+1}/{MAX_CAPTCHA_RETRIES})...")
                driver.execute_script("""
                    var links = document.querySelectorAll('a');
                    for (var i = 0; i < links.length; i++) {
                        var oc = links[i].getAttribute('onclick') || '';
                        if (oc.indexOf('loadCaptcha') > -1 || oc.indexOf('Captcha') > -1) {
                            links[i].click(); break;
                        }
                    }
                """)
                time.sleep(0.5)
                continue

            # Fill CAPTCHA answer
            captcha_input = driver.find_element(By.ID, "CaptchModel_CaptchaAnswer")
            captcha_input.clear()
            captcha_input.send_keys(answer)

            # Click View button
            view_btn = driver.find_element(By.ID, "btnView")
            view_btn.click()

            # Wait for the page to respond (results, error, or record not found)
            # instead of a fixed sleep
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: (
                        "Matching Trade Marks" in d.execute_script("return document.body.innerText.substring(0, 2000);")
                        or "Invalid Captcha" in d.execute_script("return document.body.innerText.substring(0, 500);")
                        or "invalid captcha" in d.execute_script("return document.body.innerText.substring(0, 500).toLowerCase();")
                        or "RegisteredTM" in d.current_url
                    )
                )
            except TimeoutException:
                pass  # Fall through to checks below

            page_text = driver.execute_script("return document.body.innerText.substring(0, 2000);")

            if '"code":"401"' in page_text or '"message":"Unauthorized"' in page_text:
                return "expired"

            if "Invalid Captcha" in page_text or "invalid captcha" in page_text.lower():
                print(f"    [CAPTCHA] Wrong answer, retrying ({attempt+1}/{MAX_CAPTCHA_RETRIES})...")
                continue

            # Check "Record Not Found" BEFORE "Matching Trade Marks"
            if "Record Not Found" in page_text or "No Record Found" in page_text or "no record" in page_text.lower():
                return "not_found"

            if "Matching Trade Marks" in page_text:
                return "found"

            # Check if we're on the results page by URL
            current_url = driver.current_url
            if "RegisteredTM" in current_url or "ViewRegistration" in current_url:
                if "Record Not Found" in page_text:
                    return "not_found"
                return "found"

            # Still on the form = CAPTCHA failed silently
            app_field = driver.find_elements(By.ID, "ApplicationNumber")
            if app_field:
                print(f"    [CAPTCHA] Seems to have failed silently, retrying ({attempt+1}/{MAX_CAPTCHA_RETRIES})...")
                continue

            return "found"

        except Exception as e:
            print(f"    [ERROR] Attempt {attempt+1}: {e}")
            if attempt < MAX_CAPTCHA_RETRIES - 1:
                time.sleep(0.5)
                continue
            return "error"

    return "error"


# ─── JavaScript for extracting results ──────────────────────────────────────

JS_EXTRACT_RESULTS = """
var result = {};

// Status info (above the table)
var bodyText = document.body.innerText;

// Extract Status
var statusMatch = bodyText.match(/Status:\\s*([^\\n]+)/);
result.status = statusMatch ? statusMatch[1].trim() : '';

// Extract Sub Status
var subStatusMatch = bodyText.match(/Sub Status:\\s*([^\\n]+)/);
result.subStatus = subStatusMatch ? subStatusMatch[1].trim() : '';

// Extract As on Date
var dateMatch = bodyText.match(/As on Date\\s*:\\s*([^\\n]+)/);
result.asOnDate = dateMatch ? dateMatch[1].trim() : '';

// Find the main results table
var tables = document.querySelectorAll('table');
var mainTable = null;
for (var i = 0; i < tables.length; i++) {
    if (tables[i].innerHTML.indexOf('Trade Mark No') > -1 &&
        tables[i].innerHTML.indexOf('Date of Application') > -1) {
        mainTable = tables[i];
        break;
    }
}

if (mainTable) {
    var rows = mainTable.querySelectorAll('tr');
    // First row = headers, second row = data
    if (rows.length >= 2) {
        var dataCells = rows[1].querySelectorAll('td');
        if (dataCells.length >= 10) {
            result.tradeMarkNo     = dataCells[0] ? dataCells[0].textContent.trim() : '';
            result.dateOfApp       = dataCells[1] ? dataCells[1].textContent.trim() : '';
            result.tmClass         = dataCells[2] ? dataCells[2].textContent.trim() : '';
            result.filingMode      = dataCells[3] ? dataCells[3].textContent.trim() : '';
            result.tradeMark       = dataCells[4] ? dataCells[4].textContent.trim() : '';
            result.tmType          = dataCells[5] ? dataCells[5].textContent.trim() : '';
            result.userDetail      = dataCells[6] ? dataCells[6].textContent.trim() : '';
            result.publicationDet  = dataCells[7] ? dataCells[7].textContent.trim() : '';
            result.validUpto       = dataCells[8] ? dataCells[8].textContent.trim() : '';
            result.proprietorName  = dataCells[9] ? dataCells[9].textContent.trim() : '';
        }
    }
}

// Check for Trade Mark Image — it appears below the results table
// Look for any element containing "Trade Mark Image" text, then grab the img inside/after it
var tmImageFound = false;
var allEls = document.querySelectorAll('*');
for (var k = 0; k < allEls.length; k++) {
    var elText = allEls[k].textContent.trim();
    if (elText.indexOf('Trade Mark Image') > -1 && elText.length < 60) {
        // Found the label — now find the nearest img (sibling or child)
        var parent = allEls[k].closest('div, td, fieldset, section') || allEls[k].parentElement;
        if (parent) {
            var nearbyImgs = parent.querySelectorAll('img');
            if (nearbyImgs.length > 0) {
                result.hasImage = true;
                tmImageFound = true;
                break;
            }
        }
    }
}
if (!tmImageFound) {
    result.hasImage = false;
}

return result;
"""


def extract_results(driver, app_number, country):
    """Extract trademark data from the results page."""
    record = {h: "" for h in CSV_HEADERS}
    record["Search App Number"] = app_number
    record["Country"] = country

    try:
        # Wait for the results table to appear
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            lambda d: d.find_elements(By.XPATH, "//table//th[contains(text(),'Trade Mark No')]")
                or d.find_elements(By.XPATH, "//*[contains(text(),'No Record Found')]")
        )

        # Check for "No Record Found"
        no_record = driver.find_elements(
            By.XPATH, "//*[contains(text(),'No Record Found')]"
        )
        if no_record:
            record["Status"] = "NOT FOUND ON E-REGISTER"
            return record

        # Extract data using JavaScript
        data = driver.execute_script(JS_EXTRACT_RESULTS)

        if data:
            record["Status"]                   = data.get("status", "")
            record["Sub Status"]               = data.get("subStatus", "")
            record["Trade Mark No."]           = data.get("tradeMarkNo", "")
            record["Date of Application"]      = data.get("dateOfApp", "")
            record["Class"]                    = data.get("tmClass", "")
            record["Filing Mode"]              = data.get("filingMode", "")
            record["Trade Mark"]               = data.get("tradeMark", "")
            record["TM Type"]                  = data.get("tmType", "")
            record["User Detail"]              = data.get("userDetail", "")
            record["Publication Details"]      = data.get("publicationDet", "")
            record["Valid Upto/ Renewed Upto"] = data.get("validUpto", "")
            record["Proprietor Name"]          = data.get("proprietorName", "")

            # Download the trademark image if one exists on the page
            has_image = data.get("hasImage", False)
            if has_image:
                try:
                    saved_path = download_tm_image(driver, app_number)
                    if saved_path:
                        record["Image URL"] = saved_path
                except Exception as e:
                    print(f"    [IMG] Could not download image: {e}")

    except TimeoutException:
        record["Status"] = "TIMEOUT - NO RESULTS"
    except Exception as e:
        record["Status"] = f"SCRAPE ERROR: {e}"
        print(f"    [ERROR] Extraction failed: {e}")

    return record


def download_tm_image(driver, app_number):
    """
    Find the trademark image on the page (near 'Trade Mark Image' label)
    and save it using Selenium's element screenshot.
    """
    os.makedirs(IMAGE_DIR, exist_ok=True)
    img_path = os.path.join(IMAGE_DIR, f"{app_number}.png")

    if os.path.exists(img_path):
        print(f"    [IMG] Already exists: {img_path}")
        return img_path

    try:
        # Find the trademark image element on the page
        # Strategy: locate any img element that is inside or near a container
        # with "Trade Mark Image" text
        img_element = driver.execute_script("""
            // Find elements containing "Trade Mark Image" text
            var allEls = document.querySelectorAll('*');
            for (var i = 0; i < allEls.length; i++) {
                var t = allEls[i].textContent.trim();
                if (t.indexOf('Trade Mark Image') > -1 && t.length < 60) {
                    var parent = allEls[i].closest('div, td, fieldset, section, table')
                                 || allEls[i].parentElement;
                    if (parent) {
                        var imgs = parent.querySelectorAll('img');
                        if (imgs.length > 0) return imgs[0];
                    }
                    // Also check next sibling
                    var next = allEls[i].nextElementSibling;
                    while (next) {
                        if (next.tagName === 'IMG') return next;
                        var childImgs = next.querySelectorAll('img');
                        if (childImgs.length > 0) return childImgs[0];
                        next = next.nextElementSibling;
                    }
                }
            }
            return null;
        """)

        if img_element:
            img_element.screenshot(img_path)
            print(f"    [IMG] Saved: {img_path}")
            return img_path
        else:
            print("    [IMG] No trademark image found on page")
            return None

    except Exception as e:
        print(f"    [IMG] Download error: {e}")
        return None


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  IP India E-Register Scraper — Application Number Search")
    print("=" * 65)

    tasks    = load_search_tasks()
    done_set = load_checkpoint()

    # Filter already-done entries
    remaining = []
    for (app_num, country) in tasks:
        if app_num in done_set:
            continue
        remaining.append((app_num, country))

    # Count national vs international
    national_count = sum(1 for _, c in remaining if not is_international(c))
    intl_count = sum(1 for _, c in remaining if is_international(c))

    print(f"[INFO] Total: {len(tasks)} | Done: {len(done_set)} | Remaining: {len(remaining)}")
    print(f"       National: {national_count} | International: {intl_count}")

    if not remaining:
        print("[DONE] All rows already scraped. Delete eregister_checkpoint.txt to re-run.")
        return

    init_csv()

    print("\n[BROWSER] Launching Chrome...")
    driver = init_driver()

    # Step 1: Manual OTP login
    wait_for_manual_login(driver)

    # Step 2: Navigate to search form (start with National)
    print("\n[NAV] Navigating to search form...")
    current_mode_intl = False  # Track which radio button is currently active
    if not navigate_to_search_form(driver, international=False):
        print("[ERROR] Could not navigate to search form. Exiting.")
        driver.quit()
        return

    total = len(remaining)
    scraped_count   = 0
    no_result_count = 0
    error_list      = []
    form_ready      = True  # We just navigated, so form is ready

    try:
        for idx, (app_number, country) in enumerate(remaining, 1):
            need_intl = is_international(country)
            print(f"\n{'─'*65}")
            print(f"[{idx}/{total}]  App No: {app_number}  |  Country: {country}"
                  f"  |  {'INTL' if need_intl else 'NAT'}")

            # Skip invalid application numbers (must contain digits)
            if not any(c.isdigit() for c in app_number):
                print(f"  [SKIP] '{app_number}' is not a valid application number.")
                empty = {h: "" for h in CSV_HEADERS}
                empty["Search App Number"] = app_number
                empty["Country"] = country
                empty["Status"] = "INVALID APP NUMBER"
                append_rows([empty])
                mark_done(app_number)
                no_result_count += 1
                continue

            while True:
                try:
                    # Decide how to get to the form
                    if need_intl != current_mode_intl:
                        # Mode switch — full re-navigation required
                        print(f"  [SWITCH] Changing to {'International' if need_intl else 'National'} mode...")
                        if not navigate_to_search_form(driver, international=need_intl):
                            print("  [ERROR] Could not switch search mode. Skipping.")
                            error_list.append(app_number)
                            break
                        current_mode_intl = need_intl
                        form_ready = True
                    elif not form_ready:
                        # Form not ready (first run or after error) — re-navigate
                        navigate_to_search_form(driver, international=need_intl)
                        form_ready = True
                    # else: form is already ready from Back button — skip navigation

                    # Fill form and submit
                    result = fill_and_submit(driver, app_number, international=need_intl)

                    if result == "expired":
                        handle_session_expiry(driver)
                        form_ready = False
                        continue  # Retry this application number

                    if result == "not_found":
                        print("  [NO RESULT] Not found on E-Register.")
                        empty = {h: "" for h in CSV_HEADERS}
                        empty["Search App Number"] = app_number
                        empty["Country"] = country
                        empty["Status"] = "NOT FOUND ON E-REGISTER"
                        append_rows([empty])
                        mark_done(app_number)
                        no_result_count += 1
                        # Click Back to return to form
                        form_ready = go_back_to_form(driver)
                        break

                    elif result == "error":
                        print("  [ERROR] Could not complete search.")
                        error_list.append(app_number)
                        form_ready = False  # Force re-navigation next time
                        break

                    else:  # "found"
                        print("  [OK] Results found! Extracting...")
                        record = extract_results(driver, app_number, country)
                        print(f"  [>] {record.get('Trade Mark No.', '?')}  |  "
                              f"{record.get('Status', '?')}  |  "
                              f"{record.get('Proprietor Name', '?')}")
                        append_rows([record])
                        mark_done(app_number)
                        scraped_count += 1
                        # Click Back to return to form
                        form_ready = go_back_to_form(driver)
                        break

                except KeyboardInterrupt:
                    raise
                except Exception as e:
                    print(f"  [ERROR] {e}")
                    error_list.append(app_number)
                    form_ready = False  # Force re-navigation next time
                    break

            # Pace the requests to avoid IP blocking
            human_delay()

            # Coffee break every BATCH_SIZE searches
            searches_done = scraped_count + no_result_count + len(error_list)
            if searches_done > 0 and searches_done % BATCH_SIZE == 0:
                pause = random.uniform(BATCH_BREAK_MIN, BATCH_BREAK_MAX)
                print(f"\n  ☕ Coffee break ({searches_done} searches done). "
                      f"Pausing {pause:.0f}s to avoid detection...")
                time.sleep(pause)
                print("  ☕ Resuming...")

    except KeyboardInterrupt:
        print("\n\n[INTERRUPTED] Stopped by user. Progress is saved.")

    finally:
        print(f"\n{'='*65}")
        print(f"  SUMMARY")
        print(f"  Records scraped : {scraped_count}")
        print(f"  Not found       : {no_result_count}")
        print(f"  Errors/Skipped  : {len(error_list)}")
        if error_list:
            print(f"  Error list      : {error_list[:20]}")
            if len(error_list) > 20:
                print(f"                    ... and {len(error_list)-20} more")
        print(f"  Output CSV      : {OUTPUT_CSV}")
        print(f"  Checkpoint      : {CHECKPOINT}")
        print(f"  TM Images       : {IMAGE_DIR}")
        print(f"{'='*65}")
        # Don't quit — keep browser open (detach mode)
        print("\n[INFO] Browser left open. Close it manually when done.")


if __name__ == "__main__":
    main()
