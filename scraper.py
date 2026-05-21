"""
IP India Trademark Scraper — AUTO CAPTCHA
==========================================
Scrapes trademark data from https://tmrsearch.ipindia.gov.in/tmrpublicsearch/
Solves CAPTCHAs automatically via the server-side GetCaptcha endpoint.

Usage:
    python scraper.py

Input:  Place your Excel file (.xlsx) in the project folder
        Must have columns: "Trademark Name" and "International classes"

Output: output.csv           — scraped trademark data
Resume: checkpoint.txt       — tracks progress (safe to stop & resume)
"""

import csv
import json
import os
import re
import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass
import time
import traceback

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import NoSuchElementException, TimeoutException

# ── Config ───────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH     = os.path.join(BASE_DIR, "input.xlsx")
OUTPUT_CSV     = os.path.join(BASE_DIR, "output.csv")
CHECKPOINT     = os.path.join(BASE_DIR, "checkpoint.txt")
BASE_URL       = "https://tmrsearch.ipindia.gov.in/tmrpublicsearch/"
WAIT_TIMEOUT   = 15
BETWEEN_SEARCH = 0.3
MAX_CAPTCHA_RETRIES = 5
# ─────────────────────────────────────────────────────────────────────────────

CSV_HEADERS = [
    "Search Keyword", "Search Class", "Record No", "Application Number",
    "Word Mark", "Class", "Status", "Appl. Date", "Proprietor",
    "Journal No", "Journal Date", "Used Since", "Valid Upto",
    "Goods & Services", "Address", "Agent / Attorney", "Other Details",
]


# ─── CAPTCHA Solver ─────────────────────────────────────────────────────────

def solve_captcha(driver):
    """Solve CAPTCHA via the server-side GetCaptcha endpoint."""
    try:
        result = driver.execute_async_script("""
            var cb = arguments[arguments.length - 1];
            fetch('/tmrpublicsearch/frmmain.aspx/GetCaptcha', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: '{}'
            })
            .then(r => r.text())
            .then(t => cb(t))
            .catch(e => cb('ERROR:' + e.message));
        """)
        if result and not result.startswith('ERROR'):
            data = json.loads(result)
            if isinstance(data, dict) and 'd' in data:
                text = data['d']
                if text and len(text) >= 3:
                    print(f"    [CAPTCHA] Solved: {text}")
                    return text
    except Exception as e:
        print(f"    [CAPTCHA] Endpoint error: {e}")
    return None


def fill_and_submit(driver, tm_name, class_str):
    """Fill the search form, solve CAPTCHA, and click search."""
    for attempt in range(MAX_CAPTCHA_RETRIES):
        try:
            if attempt > 0:
                driver.get(BASE_URL)
                time.sleep(0.8)

            ddl = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "ContentPlaceHolder1_DDLSearchType"))
            )
            Select(ddl).select_by_value("WM")
            time.sleep(0.1)

            wm = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.ID, "ContentPlaceHolder1_TBWordmark"))
            )
            wm.clear()
            wm.send_keys(tm_name)

            try:
                ci = driver.find_element(By.ID, "ContentPlaceHolder1_TBClass")
                ci.clear()
                if class_str:
                    first_class = class_str.replace(".", ",").split(",")[0].strip()
                    ci.send_keys(first_class)
            except Exception:
                pass

            captcha_text = solve_captcha(driver)
            if not captcha_text:
                print(f"    [CAPTCHA] Failed to get answer, retrying ({attempt+1}/{MAX_CAPTCHA_RETRIES})...")
                continue

            captcha_input = None
            for inp_id in [
                "ContentPlaceHolder1_captcha1",
                "ContentPlaceHolder1_TBCaptcha",
                "ContentPlaceHolder1_txtCaptcha",
            ]:
                els = driver.find_elements(By.ID, inp_id)
                if els:
                    captcha_input = els[0]
                    break

            if not captcha_input:
                inputs = driver.find_elements(By.XPATH,
                    "//input[@type='text' and (contains(@id,'aptcha') or contains(@id,'code') or contains(@placeholder,'code'))]"
                )
                if inputs:
                    captcha_input = inputs[0]

            if not captcha_input:
                raise RuntimeError("Could not locate CAPTCHA input field")

            captcha_input.clear()
            captcha_input.send_keys(captcha_text)
            time.sleep(0.1)

            btn = None
            for btn_id in [
                "ContentPlaceHolder1_BtnSearch",
                "ContentPlaceHolder1_btnSearch",
                "ContentPlaceHolder1_ButtonSearch",
            ]:
                els = driver.find_elements(By.ID, btn_id)
                if els:
                    btn = els[0]
                    break

            if not btn:
                btns = driver.find_elements(By.XPATH, "//input[@type='submit' or @type='button'][@value='Search']")
                if btns:
                    btn = btns[0]

            if btn:
                btn.click()
            else:
                raise RuntimeError("Could not locate Search button")

            time.sleep(0.8)

            error_els = driver.find_elements(By.XPATH,
                "//*[contains(text(),'Invalid Captcha') or contains(text(),'invalid captcha') "
                "or contains(text(),'Wrong Captcha') or contains(text(),'wrong captcha') "
                "or contains(text(),'incorrect') or contains(text(),'Incorrect')]"
            )
            if error_els and any(e.is_displayed() for e in error_els):
                print(f"    [CAPTCHA] Wrong answer, retrying ({attempt+1}/{MAX_CAPTCHA_RETRIES})...")
                refresh_btns = driver.find_elements(By.ID, "ContentPlaceHolder1_ImageButton1")
                if refresh_btns:
                    refresh_btns[0].click()
                    time.sleep(0.5)
                continue

            return True

        except Exception as e:
            print(f"    [CAPTCHA] Attempt {attempt+1} failed: {e}")
            if attempt < MAX_CAPTCHA_RETRIES - 1:
                time.sleep(1)
                continue
            return False

    return False


# ─── Excel Loading ────────────────────────────────────────────────────────────

def load_search_tasks():
    if not os.path.exists(EXCEL_PATH):
        sys.exit(f"ERROR: Excel file not found: {EXCEL_PATH}\n"
                 f"Place your .xlsx file as 'input.xlsx' in the project folder.")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        tm_col    = headers.index("Trademark Name")
        class_col = headers.index("International classes")
    except ValueError as e:
        sys.exit(f"ERROR: Required column not found in Excel: {e}")

    tasks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        tm  = row[tm_col]
        cls = row[class_col]
        if tm is None:
            continue
        tm_str  = str(tm).strip()
        cls_str = str(cls).strip() if cls is not None else ""
        if tm_str and tm_str.lower() != "none":
            tasks.append((tm_str, cls_str))

    print(f"[EXCEL] Loaded {len(tasks)} rows from {EXCEL_PATH}")
    return tasks


# ─── Checkpoint ───────────────────────────────────────────────────────────────

def make_key(tm, cls):
    return f"{tm}|||{cls}"

def load_checkpoint():
    done = set()
    if os.path.exists(CHECKPOINT):
        with open(CHECKPOINT, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    done.add(line)
    return done

def mark_done(tm, cls):
    with open(CHECKPOINT, "a", encoding="utf-8") as f:
        f.write(make_key(tm, cls) + "\n")


# ─── CSV ──────────────────────────────────────────────────────────────────────

def init_csv():
    if not os.path.exists(OUTPUT_CSV):
        with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
            csv.DictWriter(f, fieldnames=CSV_HEADERS).writeheader()
        print(f"[CSV] Created: {OUTPUT_CSV}")
    else:
        print(f"[CSV] Will append to existing: {OUTPUT_CSV}")

def append_rows(rows):
    if not rows:
        return
    with open(OUTPUT_CSV, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CSV_HEADERS, extrasaction="ignore")
        for row in rows:
            w.writerow(row)


# ─── Browser ─────────────────────────────────────────────────────────────────

def init_driver():
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-gpu")
    options.page_load_strategy = 'eager'
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    driver.set_script_timeout(15)
    driver.implicitly_wait(1)
    return driver


# ─── Wait for Results ────────────────────────────────────────────────────────

def wait_for_results(driver):
    """Returns: 'found', 'not_found', or 'timeout'"""
    start = time.time()
    while time.time() - start < WAIT_TIMEOUT:
        try:
            if driver.find_elements(By.PARTIAL_LINK_TEXT, "Show Details"):
                return "found"
            if driver.find_elements(By.ID, "ContentPlaceHolder1_MGVSearchResult_Label6_0"):
                return "found"

            lbl = driver.find_elements(By.ID, "ContentPlaceHolder1_MGVSearchResult_LblEmptyData")
            if lbl:
                return "not_found"

            no_rec = driver.find_elements(
                By.XPATH,
                "//*[@id='searchGrid' or contains(@id,'DivSearchResult')]"
                "//td[contains(text(),'No Record')]"
            )
            if no_rec and any(el.is_displayed() for el in no_rec):
                return "not_found"

            matching_hdr = driver.find_elements(
                By.XPATH, "//*[contains(text(),'Matching Trademark')]"
            )
            if matching_hdr:
                grid_rows = driver.find_elements(
                    By.XPATH,
                    "//*[contains(@id,'MGVSearchResult') and contains(@id,'lblapplicationnumber')]"
                )
                if not grid_rows:
                    time.sleep(0.4)
                    if not driver.find_elements(By.PARTIAL_LINK_TEXT, "Show Details"):
                        return "not_found"

        except Exception:
            pass
        time.sleep(0.5)
    return "timeout"


# ─── Fast JS Constants ──────────────────────────────────────────────────────

JS_CLICK_AND_READ_GRID = """
var i = arguments[0];
var p = 'ContentPlaceHolder1_MGVSearchResult';
var g = function(id) { var e = document.getElementById(id); return e ? e.textContent.trim() : ''; };
var links = document.querySelectorAll('a');
var count = 0;
for (var j = 0; j < links.length; j++) {
    if (links[j].textContent.indexOf('Show Details') > -1) {
        if (count === i) { links[j].click(); break; }
        count++;
    }
}
return {
    app: g(p + '_lblapplicationnumber_' + i),
    wm:  g(p + '_lblsimiliarmark_' + i),
    cls: g(p + '_lblsearchclass_' + i),
    st:  g(p + '_Label6_' + i)
};
"""

JS_READ_PANEL = """
var tables = document.querySelectorAll('table');
var panel = null;
for (var i = tables.length - 1; i >= 0; i--) {
    if (tables[i].innerHTML.indexOf('Appl. No.') > -1) { panel = tables[i]; break; }
}
if (!panel) return null;
function field(label) {
    var bolds = panel.querySelectorAll('b');
    for (var j = 0; j < bolds.length; j++) {
        if (bolds[j].textContent.indexOf(label) > -1) {
            var td = bolds[j].closest('td');
            if (td) {
                var next = td.nextElementSibling;
                if (next && next.tagName === 'TD') return next.textContent.trim();
                var raw = td.textContent.trim();
                var idx = raw.indexOf(label);
                if (idx > -1) return raw.substring(idx + label.length).replace(/^[\\s:]+/, '').trim();
            }
        }
    }
    return '';
}
var recNo = 0;
var allTds = panel.querySelectorAll('td');
for (var k = 0; k < allTds.length; k++) {
    var t = allTds[k].textContent;
    if (t.indexOf('Record No') > -1) {
        var m = t.match(/Record No[.\\s:]+(\\d+)/);
        if (m) { recNo = parseInt(m[1]); break; }
    }
}
return {
    recNo: recNo, applDate: field('Appl. Date'), proprietor: field('Proprietor'),
    goods: field('Goods & Services') || field('Goods \\u0026 Services'),
    address: field('Address'), agent: field('Agent'), other: field('Other Details'),
    journalNo: field('Journal No'), journalDate: field('Journal Date'),
    usedSince: field('Used Since'), validUpto: field('Valid Upto'),
    wordMark: field('Word Mark'), applNo: field('Appl. No.')
};
"""

JS_COUNT_LINKS = """
var links = document.querySelectorAll('a');
var count = 0;
for (var i = 0; i < links.length; i++) {
    if (links[i].textContent.indexOf('Show Details') > -1) count++;
}
return count;
"""


# ─── Fast Panel Wait ─────────────────────────────────────────────────────────

def wait_for_panel_js(driver, expected_record_no, app_num, timeout=4):
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            result = driver.execute_script("""
                var tables = document.querySelectorAll('table');
                for (var i = tables.length - 1; i >= 0; i--) {
                    if (tables[i].innerHTML.indexOf('Appl. No.') > -1) {
                        var text = tables[i].textContent;
                        var m = text.match(/Record No[.\\s:]+(\\d+)/);
                        var rn = m ? parseInt(m[1]) : 0;
                        return {recNo: rn, text: text.substring(0, 500)};
                    }
                }
                return null;
            """)
            if result:
                if expected_record_no > 0 and result.get('recNo') == expected_record_no:
                    return True
                if app_num and app_num in result.get('text', ''):
                    return True
        except Exception:
            pass
        time.sleep(0.08)
    return False


# ─── Load More ────────────────────────────────────────────────────────────────

def load_all_results(driver):
    clicks = 0
    while True:
        try:
            clicked = driver.execute_script("""
                var links = document.querySelectorAll('a');
                for (var i = 0; i < links.length; i++) {
                    var t = links[i].textContent.toLowerCase();
                    if (t.indexOf('load more') > -1 && links[i].offsetParent !== null) {
                        links[i].click();
                        return true;
                    }
                }
                return false;
            """)
            if not clicked:
                break
            clicks += 1
            time.sleep(0.4)
            print(f"    [LOAD MORE] Clicked (batch {clicks})", flush=True)
        except Exception:
            break
    if clicks:
        print(f"    [LOAD MORE] Loaded {clicks} extra page(s) of results.")


# ─── Main Scrape Logic ───────────────────────────────────────────────────────

def scrape_all_results(driver, tm_name, class_str):
    records = []
    load_all_results(driver)

    total_links = driver.execute_script(JS_COUNT_LINKS)
    if not total_links:
        print("    (no Show Details links found)")
        return records

    for index in range(total_links):
        try:
            grid = driver.execute_script(JS_CLICK_AND_READ_GRID, index)
            app_num   = grid.get('app', '') if grid else ''
            word_mark = grid.get('wm', '')  if grid else ''
            cls       = grid.get('cls', '') if grid else ''
            status    = grid.get('st', '')  if grid else ''

            time.sleep(0.15)
            wait_for_panel_js(driver, index + 1, app_num, timeout=4)

            pf = driver.execute_script(JS_READ_PANEL)

            record = {h: "" for h in CSV_HEADERS}
            record["Search Keyword"]     = tm_name
            record["Search Class"]       = class_str
            record["Record No"]          = str(index + 1)
            record["Application Number"] = app_num
            record["Word Mark"]          = word_mark
            record["Class"]              = cls
            record["Status"]             = status

            if pf:
                record["Appl. Date"]       = pf.get('applDate', '')
                record["Proprietor"]       = pf.get('proprietor', '')
                record["Goods & Services"] = pf.get('goods', '')
                record["Address"]          = pf.get('address', '')
                record["Agent / Attorney"] = pf.get('agent', '')
                record["Other Details"]    = pf.get('other', '')

                j_raw = pf.get('journalNo', '')
                if "Journal Date" in j_raw:
                    parts = j_raw.split("Journal Date")
                    record["Journal No"]   = parts[0].replace("Journal No.", "").replace("Journal No", "").replace(":", "").strip()
                    record["Journal Date"] = parts[1].replace(":", "").strip()
                else:
                    record["Journal No"]   = j_raw.replace("Journal No.", "").replace("Journal No", "").replace(":", "").strip()
                    record["Journal Date"] = pf.get('journalDate', '').replace("Journal Date", "").replace(":", "").strip()

                us_raw = pf.get('usedSince', '')
                if "Valid Upto" in us_raw:
                    parts = us_raw.split("Valid Upto")
                    record["Used Since"] = parts[0].replace("Used Since", "").replace(":", "").strip()
                    record["Valid Upto"] = parts[1].replace(":", "").strip()
                else:
                    record["Used Since"] = us_raw.replace("Used Since", "").replace(":", "").strip()
                    record["Valid Upto"] = pf.get('validUpto', '').replace("Valid Upto", "").replace(":", "").strip()

                if not record["Word Mark"]:
                    record["Word Mark"] = pf.get('wordMark', '')
                if not record["Application Number"]:
                    record["Application Number"] = pf.get('applNo', '')

            print(f"    [{index+1}/{total_links}] {record['Word Mark'] or '?'}  |  {record['Status'] or '?'}")
            records.append(record)

        except Exception as e:
            print(f"    [{index+1}/{total_links}] ERROR: {e}")
            err_rec = {h: "" for h in CSV_HEADERS}
            err_rec["Search Keyword"] = tm_name
            err_rec["Search Class"]   = class_str
            err_rec["Record No"]      = str(index + 1)
            err_rec["Status"]         = f"SCRAPE ERROR: {e}"
            records.append(err_rec)

    print(f"    (done — {len(records)} rows processed)")
    return records


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  IP India Trademark Scraper — AUTO CAPTCHA")
    print("=" * 65)

    tasks    = load_search_tasks()
    done_set = load_checkpoint()
    remaining = [(tm, cls) for (tm, cls) in tasks if make_key(tm, cls) not in done_set]

    print(f"[INFO] Total: {len(tasks)} | Done: {len(done_set)} | Remaining: {len(remaining)}")

    if not remaining:
        print("[DONE] All rows already scraped. Delete checkpoint.txt to re-run.")
        return

    init_csv()

    print("\n[BROWSER] Launching Chrome...")
    driver = init_driver()

    total = len(remaining)
    scraped_count   = 0
    no_result_count = 0
    error_list      = []

    try:
        for idx, (tm_name, class_str) in enumerate(remaining, 1):
            print(f"\n{'─'*65}")
            print(f"[{idx}/{total}]  {tm_name!r}  |  Class: {class_str}")

            try:
                driver.get(BASE_URL)
                time.sleep(1.5)

                ok = fill_and_submit(driver, tm_name, class_str)
                if not ok:
                    print("  [SKIP] Could not solve CAPTCHA after retries.")
                    error_list.append((tm_name, class_str))
                    continue

                print("  [WAITING] Checking for results...")
                status = wait_for_results(driver)

                if status == "not_found":
                    print("  [NO RESULT] No records on portal.")
                    empty = {h: "" for h in CSV_HEADERS}
                    empty["Search Keyword"] = tm_name
                    empty["Search Class"]   = class_str
                    empty["Status"]         = "NOT FOUND ON PORTAL"
                    append_rows([empty])
                    mark_done(tm_name, class_str)
                    no_result_count += 1

                elif status == "timeout":
                    print("  [TIMEOUT] No results. Skipping.")
                    error_list.append((tm_name, class_str))

                else:
                    print("  [OK] Results found! Scraping...")
                    records = scrape_all_results(driver, tm_name, class_str)
                    print(f"  [>] Scraped {len(records)} record(s).")
                    append_rows(records)
                    mark_done(tm_name, class_str)
                    scraped_count += len(records)

            except KeyboardInterrupt:
                raise
            except Exception as e:
                print(f"  [ERROR] {e}")
                traceback.print_exc()
                error_list.append((tm_name, class_str))
                try:
                    driver.get(BASE_URL)
                    time.sleep(1)
                except Exception:
                    pass

            time.sleep(BETWEEN_SEARCH)

    except KeyboardInterrupt:
        print("\n\n[INTERRUPTED] Stopped by user. Progress is saved.")

    finally:
        print(f"\n{'='*65}")
        print(f"  SUMMARY")
        print(f"  Records scraped : {scraped_count}")
        print(f"  Not found       : {no_result_count}")
        print(f"  Errors/Skipped  : {len(error_list)}")
        if error_list:
            print(f"  Error list      : {error_list}")
        print(f"  Output CSV      : {OUTPUT_CSV}")
        print(f"  Checkpoint      : {CHECKPOINT}")
        print(f"{'='*65}")
        try:
            driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
